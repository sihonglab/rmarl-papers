"""Step 8g: (1) add the index number (xlsx No. 147-193) to each cite-analysis
gap entry in references.bib, and (2) rename the downloaded gap PDFs in
'download papers/' to '<No>_<title>.pdf'.
"""
import os
import re
import glob
import shutil
import datetime
import openpyxl
import fitz
from rapidfuzz import fuzz, process
import common as C

BIB = os.path.join(C.PROJECT_DIR, "collected-papers", "references.bib")
DLDIR = os.path.join(C.PROJECT_DIR, "download papers")


def norm(s):
    return C.normalize_title(re.sub(r"\bmarl\b",
                                    "multi agent reinforcement learning", s, flags=re.I))


def safe(t):
    """Mimic the collected-papers/ naming: short title, 'MARL' abbreviated,
    ':' rendered as ' -' (e.g. '147_ROMAX - Certifiably Robust Deep MARL ...')."""
    t = C.fix_ligatures(t)
    t = re.sub(r"multi-?agent reinforcement learning", "MARL", t, flags=re.I)
    t = re.sub(r"\bmulti-?agent\b", "Multi-Agent", t, flags=re.I)
    t = t.replace(":", " -")
    t = re.sub(r"[\\/*?\"<>|]", "", t).replace("–", "-")
    return re.sub(r"\s+", " ", t).strip(" -")[:85]


def load_map():
    """bibkey -> (No, Title); plus arxiv-id -> bibkey and title list."""
    ws = openpyxl.load_workbook(BIB.replace("collected-papers/references.bib",
                                            "robust-marl-papers.xlsx"))["Robust MARL Papers"]
    col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    m, ax = {}, {}
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, col["No"]).value
        if not isinstance(no, int) or no < 147:
            continue
        key = ws.cell(r, col["BibKey"]).value
        title = ws.cell(r, col["Title"]).value or ""
        link = ws.cell(r, col["Link"]).value or ""
        m[key] = (no, title)
        a = re.search(r"arxiv\.org/abs/([\w.]+)", link)
        if a:
            ax[a.group(1).split("v")[0]] = key
    return m, ax


def add_index_to_bib(keymap):
    bib = open(BIB, encoding="utf-8").read()
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    shutil.copy2(BIB, BIB + f".backup-{ts}.bib")
    n = 0
    for key, (no, _) in keymap.items():
        # idempotent: strip an existing "% No.x" line right above the entry
        bib = re.sub(r"% No\.\d+\n(?=@\w+\{" + re.escape(key) + r",)", "", bib)
        bib, c = re.subn(r"(@\w+\{" + re.escape(key) + r",)",
                         f"% No.{no}\n\\1", bib, count=1)
        n += c
    open(BIB, "w", encoding="utf-8").write(bib)
    print(f"bib: indexed {n}/{len(keymap)} gap entries")


def identify(pdf, keymap, ax, gnorm, gkeys):
    """Return the bibkey a downloaded PDF belongs to."""
    stem = os.path.splitext(os.path.basename(pdf))[0]
    if stem in keymap:                                   # named <key>.pdf
        return stem
    a = re.search(r"(\d{4}\.\d{4,5})", stem)             # arXiv id in filename
    if a and a.group(1) in ax:
        return ax[a.group(1)]
    try:                                                  # fall back to title
        d = fitz.open(pdf); txt = d[0].get_text()[:800]; d.close()
    except Exception:
        txt = ""
    b = process.extractOne(norm(txt), gnorm, scorer=fuzz.token_set_ratio)
    return gkeys[b[2]] if b and b[1] >= 75 else None


def rename_pdfs(keymap, ax):
    gkeys = list(keymap)
    gnorm = [norm(keymap[k][1]) for k in gkeys]
    done, skip = 0, []
    for pdf in sorted(glob.glob(os.path.join(DLDIR, "*.pdf"))):
        key = identify(pdf, keymap, ax, gnorm, gkeys)
        if not key or key not in keymap:
            skip.append(os.path.basename(pdf)); continue
        no, title = keymap[key]
        newname = f"{no}_{safe(title)}.pdf"
        newpath = os.path.join(DLDIR, newname)
        if os.path.abspath(pdf) == os.path.abspath(newpath):
            done += 1; continue
        if os.path.exists(newpath):
            os.remove(pdf); done += 1; continue           # already renamed copy
        os.rename(pdf, newpath)
        print(f"  {os.path.basename(pdf)[:40]:40} -> {newname}")
        done += 1
    print(f"renamed/ok {done}; unidentified: {skip}")


def main():
    keymap, ax = load_map()
    add_index_to_bib(keymap)
    rename_pdfs(keymap, ax)


if __name__ == "__main__":
    main()
