"""Step 8e: (1) replace the 3 wrong gap entries in references.bib with the
correct papers, then (2) sync the supplement cite keys in robust-marl-MASTER.md
and robust-marl-papers.xlsx to match the bib keys, so every \\cite resolves.

The 3 corrected entries are read from /tmp/fix3.json (built in the fetch step).
Backs up each file it edits. Idempotent.
"""
import os
import re
import json
import shutil
import datetime
import openpyxl
from rapidfuzz import fuzz, process
import common as C

BIB = os.path.join(C.PROJECT_DIR, "collected-papers", "references.bib")
MD = os.path.join(C.PROJECT_DIR, "robust-marl-MASTER.md")
XLSX = os.path.join(C.PROJECT_DIR, "robust-marl-papers.xlsx")
WRONG = {"mazurowski2020emergence": "blumenkamp2020emergence",
         "abul2024resilient": "ye2024resilient",
         "gao2020fault": "yang2020fault"}


def norm(s):
    return C.normalize_title(re.sub(r"\bmarl\b",
                                    "multi agent reinforcement learning", s, flags=re.I))


def pretty(b):
    m = re.match(r"@(\w+)\s*\{([^,]+),(.*)\}\s*$", b.strip(), re.S)
    typ, key, body = m.group(1), m.group(2).strip(), m.group(3).strip().rstrip(",")
    fields = re.split(r",\s*(?=[A-Za-z]+\s*=)", body)
    lines = ",\n".join("  " + f.strip() for f in fields if f.strip())
    return f"@{typ}{{{key},\n{lines}\n}}"


def backup(path):
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    b = f"{path}.backup-{ts}" + os.path.splitext(path)[1]
    shutil.copy2(path, b)
    return os.path.basename(b)


def fix_bib(fix3):
    text = open(BIB, encoding="utf-8").read()
    n = 0
    for oldkey, newkey in WRONG.items():
        pat = re.compile(r"(?:%[^\n]*\n)?@\w+\s*\{\s*" + re.escape(oldkey) +
                         r"\s*,.*?\n\}\n?", re.S)
        repl = pretty(fix3[newkey]) + "\n"
        text, c = pat.subn(repl, text)
        n += c
        print(f"  bib: {oldkey} -> {newkey}  ({'replaced' if c else 'NOT FOUND'})")
    open(BIB, "w", encoding="utf-8").write(text)
    return n


def gap_entries():
    """(key, normalized title) for the 47 entries in the gap block."""
    block = open(BIB, encoding="utf-8").read().split("BEGIN cite-analysis")[1]
    out = []
    for m in re.finditer(r"@(\w+)\s*\{\s*([^,]+),(.*?)\n\}", block, re.S):
        t = re.search(r"title\s*=\s*[{\"](.+?)[}\"]\s*,", m.group(3), re.S | re.I)
        out.append((m.group(2).strip(),
                    re.sub(r"[{}]", "", t.group(1)) if t else ""))
    return out


def sync_md(gaps):
    gnorm = [norm(t) for _, t in gaps]
    keys = {k for k, _ in gaps}
    text = open(MD, encoding="utf-8").read()
    pre, rest = text.split("BEGIN cite-analysis-supplement", 1)
    sup, post = rest.split("END cite-analysis", 1)
    changed = 0
    out_lines = []
    for line in sup.splitlines():
        mk = re.search(r"`([A-Za-z][A-Za-z0-9]+)`", line)
        if mk and mk.group(1) not in keys:
            mt = re.search(r"^\d+\.\s+(?:\[([^\]]+)\]|([^`]+?)\s+`)", line)
            title = (mt.group(1) or mt.group(2)).strip() if mt else ""
            b = process.extractOne(norm(title), gnorm, scorer=fuzz.token_sort_ratio)
            if b and b[1] >= 83:
                newk = gaps[b[2]][0]
                if newk != mk.group(1):
                    line = line.replace(f"`{mk.group(1)}`", f"`{newk}`")
                    changed += 1
        out_lines.append(line)
    new = (pre + "BEGIN cite-analysis-supplement" + "\n".join(out_lines) +
           "END cite-analysis" + post)
    open(MD, "w", encoding="utf-8").write(new)
    return changed


def sync_xlsx(gaps):
    gnorm = [norm(t) for _, t in gaps]
    keys = {k for k, _ in gaps}
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Robust MARL Papers"]
    col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    changed = 0
    for r in range(2, ws.max_row + 1):
        no = ws.cell(r, col["No"]).value
        if not isinstance(no, int) or no < 147:
            continue
        bibk = ws.cell(r, col["BibKey"]).value
        if bibk in keys:
            continue
        title = ws.cell(r, col["Title"]).value or ""
        b = process.extractOne(norm(title), gnorm, scorer=fuzz.token_sort_ratio)
        if b and b[1] >= 83 and gaps[b[2]][0] != bibk:
            ws.cell(r, col["BibKey"]).value = gaps[b[2]][0]
            changed += 1
    wb.save(XLSX)
    return changed


def main():
    fix3 = json.load(open("/tmp/fix3.json"))
    print("backups:", backup(BIB), backup(MD), backup(XLSX))
    print("1) fixing wrong bib entries:")
    fix_bib(fix3)
    gaps = gap_entries()
    print(f"2) syncing keys against {len(gaps)} gap bib entries")
    print(f"   master.md  : {sync_md(gaps)} keys updated")
    print(f"   xlsx       : {sync_xlsx(gaps)} keys updated")


if __name__ == "__main__":
    main()
