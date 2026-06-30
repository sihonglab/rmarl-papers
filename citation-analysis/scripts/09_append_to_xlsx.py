"""Step 8b: append the prepared gap rows (data/xlsx_new_rows.json) to
robust-marl-papers.xlsx, preserving existing formatting and the second sheet.
A timestamped backup is written first. Re-running is safe: rows previously
added by this tool (Source == 'cite-analysis') are removed before re-appending.
"""
import os
import shutil
import datetime
import openpyxl
import common as C

XLSX = os.path.join(C.PROJECT_DIR, "robust-marl-papers.xlsx")
SHEET = "Robust MARL Papers"
SOURCE_TAG = "cite-analysis"
# column index (1-based) per header
COL = {"No": 1, "Category": 2, "Title": 3, "Venue/Authors": 4, "Year": 5,
       "Priority": 6, "Source": 7, "BibKey": 8, "Notes/Cross-ref": 9, "Link": 10}


def main():
    rows = C.load_json("xlsx_new_rows.json")
    ts = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = os.path.join(C.PROJECT_DIR, f"robust-marl-papers.backup-{ts}.xlsx")
    shutil.copy2(XLSX, backup)

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    # remove any rows previously added by this tool (idempotent re-runs)
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, COL["Source"]).value == SOURCE_TAG:
            ws.delete_rows(r, 1)

    existing_nos = [ws.cell(r, COL["No"]).value for r in range(2, ws.max_row + 1)]
    existing_nos = [n for n in existing_nos if isinstance(n, int)]
    next_no = max(existing_nos) + 1
    existing_keys = {str(ws.cell(r, COL["BibKey"]).value).lower()
                     for r in range(2, ws.max_row + 1)}

    added = 0
    for rec in rows:
        r = ws.max_row + 1
        bib = rec["bibkey"]
        while bib.lower() in existing_keys:        # avoid bibkey collision
            bib += "x"
        existing_keys.add(bib.lower())
        citing = ", ".join("#" + str(p) for p in rec["citing"])
        note = (f"cited by {rec['count']}× in set ({citing}); "
                f"auto-added from reference analysis")
        vals = {
            "No": next_no,
            "Category": rec["category"],
            "Title": rec["title"],
            "Venue/Authors": rec["venue_authors"] or None,
            "Year": rec["year"],
            "Priority": "high" if rec["count"] >= 5 else None,
            "Source": SOURCE_TAG,
            "BibKey": bib,
            "Notes/Cross-ref": note,
            "Link": rec["link"] or None,
        }
        for name, col in COL.items():
            ws.cell(r, col).value = vals[name]
        next_no += 1
        added += 1

    wb.save(XLSX)
    print(f"backup  : {os.path.basename(backup)}")
    print(f"appended: {added} rows (No {next_no - added}..{next_no - 1})")
    print(f"sheet now has {ws.max_row - 1} data rows")


if __name__ == "__main__":
    main()
